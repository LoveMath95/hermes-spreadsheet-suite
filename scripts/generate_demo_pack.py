#!/usr/bin/env python3
from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = REPO_ROOT / "output" / "spreadsheet" / "hermes-demo-pack"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
THIN_GRAY = Side(style="thin", color="D0D7DE")
BOX_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


@dataclass(frozen=True)
class SheetSpec:
    name: str
    headers: Sequence[str]
    rows: Sequence[Sequence[object]]
    currency_columns: Sequence[int] = ()
    percent_columns: Sequence[int] = ()
    date_columns: Sequence[int] = ()
    notes: Sequence[str] = ()


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)


def style_data_sheet(ws, spec: SheetSpec) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(spec.headers))}{len(spec.rows) + 1}"

    for index, header in enumerate(spec.headers, start=1):
        cell = ws.cell(row=1, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX_BORDER

    for row_index, row in enumerate(spec.rows, start=2):
        for column_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.border = BOX_BORDER
            cell.alignment = Alignment(vertical="top")

            if column_index in spec.currency_columns and isinstance(value, (int, float)):
                cell.number_format = '$#,##0.00;[Red]($#,##0.00)'
            elif column_index in spec.percent_columns and isinstance(value, (int, float)):
                cell.number_format = '0.0%'
            elif column_index in spec.date_columns:
                cell.number_format = 'yyyy-mm-dd'

    note_row = len(spec.rows) + 3
    if spec.notes:
        ws.cell(row=note_row, column=1, value="Demo ideas").fill = SUBHEADER_FILL
        ws.cell(row=note_row, column=1).font = Font(bold=True)
        for offset, note in enumerate(spec.notes, start=1):
            ws.cell(row=note_row + offset, column=1, value=f"• {note}")

    autosize_columns(ws)


def add_formula_debug_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Formula_Debug")
    headers = ["Date", "Product", "Region", "Units", "Revenue", "Margin %"]
    rows = [
        ["2026-04-01", "USB-C Cable", "north", 120, 1800, 0.29],
        ["2026-04-02", "USB-C Cable", "south", 95, 1425, 0.26],
        ["2026-04-03", "Docking Station", "north", 44, 6600, 0.41],
        ["2026-04-04", "Docking Station", "west", 31, 4650, 0.39],
        ["2026-04-05", "HDMI Adapter", "north", 150, 2250, 0.22],
        ["2026-04-06", "HDMI Adapter", "east", 132, 1980, 0.20],
        ["2026-04-07", "Noise-Cancel Headset", "north", 28, 5040, 0.47],
        ["2026-04-08", "Noise-Cancel Headset", "south", 18, 3240, 0.44],
    ]
    spec = SheetSpec(
        name="Formula_Debug",
        headers=headers,
        rows=rows,
        currency_columns=(5,),
        percent_columns=(6,),
        notes=(
            "Cell H11 intentionally contains a broken SUMIF formula.",
            "Prompt: why is this formula broken? Fix it and apply the corrected formula to H11.",
            "Prompt: sumif revenue of region north",
        ),
    )
    style_data_sheet(ws, spec)
    for row_index in range(2, 10):
        ws.cell(row=row_index, column=1).number_format = "yyyy-mm-dd"

    ws["H8"] = "Broken formula target"
    ws["H8"].fill = SUBHEADER_FILL
    ws["H8"].font = Font(bold=True)
    ws["H9"] = "Expected idea: sum Revenue where Region = north"
    ws["H10"] = "Current formula in H11 uses Product instead of Region"
    ws["H11"] = '=SUMIF(B:B,"north",E:E)'
    ws["H11"].number_format = '$#,##0.00;[Red]($#,##0.00)'
    ws["I11"] = "Should use column C as the criteria range"
    autosize_columns(ws)


def add_readme_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Hermes Demo Pack"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Sheet"
    ws["B3"] = "What to demo"
    ws["C3"] = "Prompt idea"
    for cell in ("A3", "B3", "C3"):
        ws[cell].fill = HEADER_FILL
        ws[cell].font = HEADER_FONT
        ws[cell].border = BOX_BORDER

    rows = [
        ("Messy_Sales", "Cleanup + normalize + dashboard", "Clean this table, standardize dates/currency, remove duplicates, then create a revenue chart on a new sheet."),
        ("Budget_Variance", "Variance analysis", "Add variance and variance % columns, then create a chart comparing Plan vs Actual."),
        ("Marketing_ROAS", "ROAS summary", "Summarize spend efficiency by channel and create a chart for Revenue vs Spend."),
        ("Inventory_Restock", "Ops workflow", "Highlight low-stock SKUs, add a reorder flag, and create a filtered urgent restock sheet."),
        ("Support_Tickets", "Grouping + summary", "Group by priority and category, summarize counts, and create a management summary sheet."),
        ("Formula_Debug", "Explain + fix formula", "Why is the formula in H11 broken? Fix it and apply the corrected formula to H11."),
        ("Image_Import_Target", "Image -> table import", "Extract the attached image into a new sheet or paste it into Image_Import_Target starting at A1."),
    ]

    for row_index, row in enumerate(rows, start=4):
        for column_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=value)
            cell.border = BOX_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72


def build_sheet_specs() -> list[SheetSpec]:
    return [
        SheetSpec(
            name="Messy_Sales",
            headers=[
                "Order ID", "Order Date", "Customer", "Region", "Rep",
                "Product", "Units", "Unit Price", "Revenue", "Status"
            ],
            rows=[
                ["SO-1001", "4/1/26", "Acme Holdings", "North", "Liam", "USB-C Cable", 120, 15, 1800, "Closed Won"],
                ["SO-1002", "2026-04-02", "Acme Holdings ", "north", "Liam", "USB-C Cable", 120, "$15.00", "1,800", "Closed Won"],
                ["SO-1003", "03-APR-2026", "Bright Retail", " SOUTH ", "Ava", "Docking Station", 44, 150, 6600, "closed won"],
                ["SO-1004", "2026/04/04", "Nova Labs", "West", "Mia", "HDMI Adapter", 150, 15, "$2,250.00", "Open"],
                ["SO-1005", "April 5 2026", "Orion Foods", "East", "Noah", "Noise-Cancel Headset", 28, 180, 5040, "Closed Won"],
                ["SO-1006", "2026-04-06", "Orion Foods", "EAST", "Noah", "Noise-Cancel Headset", 28, "180.00", 5040, "Closed won"],
                ["SO-1007", "2026-04-07", "Peak Supply", "North", "Liam", "Docking Station", 31, 150, 4650, "Closed Lost"],
                ["SO-1008", "2026-04-08", "Zenith Hotels", "south", "Ava", "USB-C Cable", 95, 15, 1425, "Closed Won"],
                ["SO-1009", "2026-04-09", "Zenith Hotels", "South", "Ava", "USB-C Cable", 95, 15, 1425, "Closed Won"],
                ["SO-1010", "2026-04-10", "Terra Retail", "West", "Mia", "Docking Station", 18, 150, 2700, "Open "],
                ["SO-1011", "2026-04-11", "Bluebird Care", "North", "Liam", "Noise-Cancel Headset", 16, 180, 2880, "Closed Won"],
                ["SO-1012", "2026-04-12", "Bluebird Care", "North ", "Liam", "Noise-Cancel Headset", 16, "$180", "$2,880", "Closed Won"],
            ],
            notes=(
                "Good for cleanup, dedupe, date normalization, currency formatting, and chart generation.",
                "Prompt: clean this table, standardize dates/currency, remove duplicates, then create a dashboard sheet.",
            ),
        ),
        SheetSpec(
            name="Budget_Variance",
            headers=["Month", "Department", "Planned Spend", "Actual Spend", "Owner"],
            rows=[
                ["2026-01", "Marketing", 45000, 47200, "Lan"],
                ["2026-02", "Marketing", 47000, 45550, "Lan"],
                ["2026-03", "Marketing", 48000, 51980, "Lan"],
                ["2026-01", "Sales", 52000, 49800, "Minh"],
                ["2026-02", "Sales", 53000, 54120, "Minh"],
                ["2026-03", "Sales", 54500, 56240, "Minh"],
                ["2026-01", "Support", 28000, 27450, "Trinh"],
                ["2026-02", "Support", 28500, 29810, "Trinh"],
                ["2026-03", "Support", 29000, 30220, "Trinh"],
            ],
            currency_columns=(3, 4),
            notes=(
                "Good for variance %, conditional formatting, and plan-vs-actual charts.",
            ),
        ),
        SheetSpec(
            name="Marketing_ROAS",
            headers=["Week", "Channel", "Campaign", "Spend", "Clicks", "Leads", "Conversions", "Revenue"],
            rows=[
                ["2026-W14", "Meta", "Spring Promo", 6200, 4180, 210, 44, 22800],
                ["2026-W14", "Google Search", "Brand Core", 5400, 2890, 184, 52, 26400],
                ["2026-W14", "TikTok", "UGC Burst", 3100, 5020, 136, 18, 6900],
                ["2026-W15", "Meta", "Retargeting", 4700, 3320, 174, 39, 20100],
                ["2026-W15", "Google Search", "High Intent", 5900, 2740, 192, 57, 30150],
                ["2026-W15", "LinkedIn", "B2B Pipeline", 4200, 980, 76, 14, 12600],
                ["2026-W16", "Meta", "Video Proof", 5100, 3595, 165, 36, 19440],
                ["2026-W16", "Google Search", "Competitor Terms", 4300, 2045, 128, 29, 15420],
                ["2026-W16", "Email", "Winback Flow", 1200, 740, 88, 21, 11880],
            ],
            currency_columns=(4, 8),
            notes=(
                "Good for ROAS analysis, channel ranking, and Revenue vs Spend chart.",
            ),
        ),
        SheetSpec(
            name="Inventory_Restock",
            headers=["SKU", "Product", "Warehouse", "On Hand", "Reorder Level", "Lead Days", "Unit Cost", "Supplier"],
            rows=[
                ["INV-1001", "USB-C Cable", "HCM", 420, 250, 7, 4.2, "CablePro"],
                ["INV-1002", "Docking Station", "HCM", 36, 60, 18, 82.0, "DockLab"],
                ["INV-1003", "HDMI Adapter", "HN", 58, 80, 9, 6.5, "SignalWorks"],
                ["INV-1004", "Noise-Cancel Headset", "HN", 24, 45, 21, 96.0, "QuietWave"],
                ["INV-1005", "USB Hub", "DN", 67, 70, 12, 18.0, "HubForge"],
                ["INV-1006", "Webcam Pro", "DN", 19, 35, 14, 74.0, "VisionGrid"],
                ["INV-1007", "Laptop Stand", "HCM", 83, 40, 10, 21.0, "DeskNest"],
                ["INV-1008", "Portable SSD", "HN", 12, 30, 16, 88.0, "ByteTrail"],
            ],
            currency_columns=(7,),
            notes=(
                "Good for reorder flags, sort/filter, low-stock highlighting, and urgent restock sheet.",
            ),
        ),
        SheetSpec(
            name="Support_Tickets",
            headers=["Ticket ID", "Created At", "Team", "Priority", "Category", "Resolution Hours", "SLA Breached", "CSAT"],
            rows=[
                ["T-4101", "2026-04-01 08:15", "L1", "High", "Billing", 6.5, "No", 4.2],
                ["T-4102", "2026-04-01 09:10", "L2", "Critical", "API", 11.3, "Yes", 3.4],
                ["T-4103", "2026-04-01 10:05", "L1", "Medium", "Login", 2.2, "No", 4.7],
                ["T-4104", "2026-04-02 11:20", "L2", "High", "Integration", 9.1, "Yes", 3.9],
                ["T-4105", "2026-04-02 13:40", "L1", "Low", "Feature Request", 18.4, "No", 4.8],
                ["T-4106", "2026-04-03 08:55", "L3", "Critical", "Outage", 5.8, "No", 4.1],
                ["T-4107", "2026-04-03 15:25", "L2", "Medium", "Reporting", 7.4, "No", 4.0],
                ["T-4108", "2026-04-04 16:05", "L1", "High", "Billing", 12.6, "Yes", 3.2],
                ["T-4109", "2026-04-04 17:30", "L3", "Critical", "API", 4.4, "No", 4.5],
            ],
            notes=(
                "Good for group-by summary, priority/category pivot, and SLA anomaly callouts.",
            ),
        ),
    ]


def draw_table_image(
    output_path: Path,
    title: str,
    subtitle: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
    footer: str,
) -> None:
    width = 1600
    height = 980
    image = Image.new("RGB", (width, height), "#F4EFE6")
    draw = ImageDraw.Draw(image)

    try:
        title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 42)
        subtitle_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 24)
        body_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 22)
        header_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 22)
        footer_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 20)
    except OSError:
        title_font = subtitle_font = body_font = header_font = footer_font = ImageFont.load_default()

    draw.rounded_rectangle((48, 40, width - 48, height - 40), radius=28, fill="#FFFDF9", outline="#CABFAF", width=3)
    draw.text((84, 78), title, fill="#202020", font=title_font)
    draw.text((84, 136), subtitle, fill="#5B5B5B", font=subtitle_font)

    left = 84
    top = 205
    table_width = width - 168
    row_height = 56
    column_width = table_width // len(headers)

    draw.rounded_rectangle((left, top, left + table_width, top + row_height), radius=16, fill="#204E7A")
    for column_index, header in enumerate(headers):
        x = left + column_index * column_width + 14
        draw.text((x, top + 14), str(header), fill="#FFFFFF", font=header_font)

    current_top = top + row_height
    for row in rows:
        draw.rectangle((left, current_top, left + table_width, current_top + row_height), outline="#D9D2C8", width=1, fill="#FFFFFF")
        for column_index, value in enumerate(row):
            x = left + column_index * column_width + 14
            draw.text((x, current_top + 14), str(value), fill="#1F1F1F", font=body_font)
        current_top += row_height

    for column_index in range(1, len(headers)):
        x = left + column_index * column_width
        draw.line((x, top, x, current_top), fill="#D9D2C8", width=1)

    draw.text((84, height - 108), footer, fill="#666666", font=footer_font)

    rotated = image.rotate(-1.4, resample=Image.Resampling.BICUBIC, expand=True, fillcolor="#F4EFE6")
    rotated.save(output_path)


def create_image_fixtures() -> None:
    sales_image_headers = ["Date", "Region", "Rep", "Product", "Units", "Revenue"]
    sales_image_rows = [
        ["2026-04-18", "North", "Liam", "Docking Station", "18", "$2,700"],
        ["2026-04-18", "South", "Ava", "USB-C Cable", "95", "$1,425"],
        ["2026-04-19", "West", "Mia", "HDMI Adapter", "150", "$2,250"],
        ["2026-04-19", "East", "Noah", "Headset", "28", "$5,040"],
        ["2026-04-20", "North", "Liam", "Headset", "16", "$2,880"],
    ]
    draw_table_image(
        OUTPUT_DIR / "image_sales_snapshot.png",
        "Northwind Retail - Regional Sales Snapshot",
        "Scanned on phone camera • 2026-04-20 • Use for image-to-sheet extraction demo",
        sales_image_headers,
        sales_image_rows,
        "Prompt: Extract this table and paste it into Image_Import_Target starting at A1.",
    )
    write_csv(OUTPUT_DIR / "image_sales_snapshot_source.csv", sales_image_headers, sales_image_rows)

    expense_headers = ["Item", "Qty", "Unit Price", "Line Total"]
    expense_rows = [
        ["Standing desk frame", "2", "$249.00", "$498.00"],
        ["Ergonomic chair", "3", "$189.00", "$567.00"],
        ["Monitor arm", "4", "$79.00", "$316.00"],
        ["Cable tray", "6", "$24.00", "$144.00"],
        ["Shipping", "1", "$85.00", "$85.00"],
    ]
    draw_table_image(
        OUTPUT_DIR / "image_office_invoice.png",
        "OfficeFit Supplies - Invoice 88431",
        "AP capture test • line-item extraction • includes totals and invoice-style layout",
        expense_headers,
        expense_rows,
        "Prompt: Extract line items from this invoice image into a new sheet named AP_Import.",
    )
    write_csv(OUTPUT_DIR / "image_office_invoice_source.csv", expense_headers, expense_rows)


def create_workbook() -> Path:
    workbook = Workbook()
    add_readme_sheet(workbook)

    for spec in build_sheet_specs():
        ws = workbook.create_sheet(spec.name)
        style_data_sheet(ws, spec)

    add_formula_debug_sheet(workbook)
    workbook.create_sheet("Image_Import_Target")

    output_path = OUTPUT_DIR / "hermes_demo_pack.xlsx"
    workbook.save(output_path)
    return output_path


def create_csv_exports() -> None:
    for spec in build_sheet_specs():
        write_csv(OUTPUT_DIR / f"{spec.name.lower()}.csv", spec.headers, spec.rows)

    formula_headers = ["Date", "Product", "Region", "Units", "Revenue", "Margin %"]
    formula_rows = [
        ["2026-04-01", "USB-C Cable", "north", 120, 1800, 0.29],
        ["2026-04-02", "USB-C Cable", "south", 95, 1425, 0.26],
        ["2026-04-03", "Docking Station", "north", 44, 6600, 0.41],
        ["2026-04-04", "Docking Station", "west", 31, 4650, 0.39],
        ["2026-04-05", "HDMI Adapter", "north", 150, 2250, 0.22],
        ["2026-04-06", "HDMI Adapter", "east", 132, 1980, 0.20],
        ["2026-04-07", "Noise-Cancel Headset", "north", 28, 5040, 0.47],
        ["2026-04-08", "Noise-Cancel Headset", "south", 18, 3240, 0.44],
    ]
    write_csv(OUTPUT_DIR / "formula_debug.csv", formula_headers, formula_rows)


def write_prompt_guide() -> None:
    content = """# Hermes Demo Prompt Pack

## Best on-camera flows

1. `Messy_Sales`
   - `Clean this table, standardize dates/currency, remove duplicates, then create a dashboard sheet with a revenue chart.`
   - `Show me the top 5 customers by revenue after cleanup.`

2. `Budget_Variance`
   - `Add variance and variance percent columns, highlight overspend rows, then create a plan vs actual chart.`

3. `Marketing_ROAS`
   - `Summarize ROAS by channel, sort from best to worst, and create a revenue vs spend chart.`

4. `Inventory_Restock`
   - `Flag low-stock items, add a reorder status column, and create a filtered urgent restock sheet.`

5. `Support_Tickets`
   - `Create a management summary of ticket counts by priority and category, then highlight SLA risk areas.`

6. `Formula_Debug`
   - `Why is the formula in H11 broken? Fix it and apply the corrected formula to H11.`
   - `sumif revenue of region north`

7. `Image fixtures`
   - Upload `image_sales_snapshot.png`
     - `Extract this table and paste it into Image_Import_Target starting at A1.`
   - Upload `image_office_invoice.png`
     - `Extract line items from this invoice image into a new sheet named AP_Import.`
"""
    (OUTPUT_DIR / "DEMO_PROMPTS.md").write_text(content, encoding="utf-8")


def main() -> None:
    ensure_output_dir()
    create_csv_exports()
    workbook_path = create_workbook()
    create_image_fixtures()
    write_prompt_guide()
    print(f"Generated demo pack in {OUTPUT_DIR}")
    print(f"Workbook: {workbook_path}")


if __name__ == "__main__":
    main()
